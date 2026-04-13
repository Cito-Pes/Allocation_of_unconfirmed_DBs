"""
미체결DB 배정 프로그램
개발환경: Python 3.13 / PySide6
DB: SQL Server 2008 R2 (pytds - 순수 Python TDS 구현)
"""

import os
import re
import sys
import sqlite3
import datetime
import requests
import pytds


from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QMessageBox,
    QLabel, QHeaderView, QCheckBox, QAbstractItemView, QSizePolicy,
    QTextEdit, QFrame, QSplitter
)
from PySide6.QtCore import Qt, QThread, Signal, QDateTime
from PySide6.QtGui import QIcon, QColor, QFont, QFontDatabase

import openpyxl
from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
#  경로 헬퍼 (PyInstaller exe / 일반 실행 공용)
# ─────────────────────────────────────────────
def app_dir() -> str:
    """실행 파일(또는 스크립트)이 있는 폴더 — Downloads·DB 저장 경로에 사용"""
    if hasattr(sys, "frozen"):          # PyInstaller exe
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def resource_path(relative: str) -> str:
    """번들 내 리소스 경로 — 폰트·아이콘 등 읽기 전용 파일에 사용"""
    if hasattr(sys, "_MEIPASS"):        # PyInstaller 압축 해제 임시 폴더
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative)


# ─────────────────────────────────────────────
#  Config DB 설정 (Google_Drive_ConfigDB_Guide.md 기준)
# ─────────────────────────────────────────────
FILE_ID = "1oncya1uYDnbVS2KwuBAKw4x4o9oQDct0"
GDRIVE_URL  = f"https://drive.google.com/file/d/{FILE_ID}/view?usp=drive_link"   # ★ 수정
# url = "https://drive.google.com/file/d/1oncya1uYDnbVS2KwuBAKw4x4o9oQDct0/view?usp=drive_link"
DB_DIR      = os.path.join(app_dir(), "DB")   # 실행파일 옆 DB 폴더
DB_FILE     = "Config_DB.db"
CONFIG_NAME = "HD_MSSQL"


def download_db() -> tuple[bool, str]:
    db_path = os.path.join(DB_DIR, DB_FILE)
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR)
    try:
        match = re.search(r"/d/([a-zA-Z0-9_-]+)", GDRIVE_URL)
        if not match:
            raise ValueError("Google Drive 파일 ID를 추출할 수 없습니다.")
        file_id = match.group(1)
        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
        session = requests.Session()
        resp = session.get(download_url, stream=True)
        resp.raise_for_status()
        if "text/html" in resp.headers.get("Content-Type", ""):
            for key, value in resp.cookies.items():
                if key.startswith("download_warning"):
                    download_url = (
                        f"https://drive.google.com/uc"
                        f"?export=download&confirm={value}&id={file_id}"
                    )
                    resp = session.get(download_url, stream=True)
                    resp.raise_for_status()
                    break
        with open(db_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True, db_path
    except Exception as e:
        return False, str(e)


def load_db_config() -> dict:
    db_path = os.path.join(DB_DIR, DB_FILE)
    try:
        conn   = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DB_Type, Host, Port, DB_Name, DB_ID, DB_PW "
            "FROM DBCON WHERE Name = ?",
            (CONFIG_NAME,)
        )
        row = cursor.fetchone()
        conn.close()
        if not row:
            raise LookupError(f"DBCON 테이블에 Name='{CONFIG_NAME}' 레코드가 없습니다.")
        return {
            "DB_Type": row[0],
            "Host"   : row[1],
            "Port"   : row[2],
            "DB_Name": row[3],
            "DB_ID"  : row[4],
            "DB_PW"  : row[5],
        }
    except Exception as e:
        raise Exception(f"DB 설정 로드 실패: {e}")


def get_mssql_connection(cfg: dict):
    """
    pytds (python-tds) — 순수 Python TDS 구현체.
    시스템 ODBC 드라이버 / FreeTDS 없이 SQL Server에 직접 접속한다.
    SQL Server 2005 이상 호환.
    """
    return pytds.connect(
        dsn=cfg["Host"],
        port=int(cfg["Port"]),
        user=cfg["DB_ID"],
        password=cfg["DB_PW"],
        database=cfg["DB_Name"],
        login_timeout=10,
        autocommit=False,
    )


# ─────────────────────────────────────────────
#  백그라운드 워커 스레드
# ─────────────────────────────────────────────
class Worker(QThread):
    result  = Signal(object)
    error   = Signal(str)
    log_msg = Signal(str)

    def __init__(self, task, *args, **kwargs):
        super().__init__()
        self.task   = task
        self.args   = args
        self.kwargs = kwargs

    def run(self):
        try:
            res = self.task(*self.args, **self.kwargs)
            self.result.emit(res)
        except Exception as e:
            self.error.emit(str(e))


# ─────────────────────────────────────────────
#  메인 윈도우
# ─────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db_config   = None
        self.detail_data = []   # DB 내용 테이블 raw data
        self.assign_data = []   # 배정 테이블 raw data
        self._input_done = False  # DB 입력 완료 플래그 (중복 실행 방지)

        self.setWindowTitle("미체결DB 배정")
        self._set_icon()
        self.resize(1600, 860)

        self._build_ui()
        self._apply_style()
        self.load_config()

    # ── 아이콘 ──────────────────────────────
    def _set_icon(self):
        for rel in ["images/app_icon.png", "icon.ico"]:
            path = resource_path(rel)
            if os.path.exists(path):
                self.setWindowIcon(QIcon(path))
                break

    # ── UI 구성 ─────────────────────────────
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(8, 8, 8, 8)
        root_layout.setSpacing(6)

        # ── 상단 메인 영역 ──────────────────
        top_splitter = QSplitter(Qt.Horizontal)

        # 1) 월별 DB 수량 테이블
        left_frame = self._make_frame("월별 DB 수량")
        self.tbl_monthly = self._make_table(["", "월", "수량"])
        self.tbl_monthly.setColumnWidth(0, 30)
        self.tbl_monthly.setColumnWidth(1, 80)
        self.tbl_monthly.setColumnWidth(2, 60)
        self.tbl_monthly.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        left_frame.layout().addWidget(self.tbl_monthly)
        top_splitter.addWidget(left_frame)

        # 2) DB 내용 테이블
        mid_frame = self._make_frame("DB 내용")
        self.tbl_detail = self._make_table([
            "휴대전화", "ID", "성명", "등록일", "상태",
            "사번", "담당자명", "권", "배정일",
            "주문번호", "채널", "신규배정일", "신규담당자"
        ])
        self.tbl_detail.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        mid_frame.layout().addWidget(self.tbl_detail)
        top_splitter.addWidget(mid_frame)

        top_splitter.setSizes([220, 1100])
        root_layout.addWidget(top_splitter, stretch=6)

        # ── 하단 영역 ───────────────────────
        bottom_splitter = QSplitter(Qt.Horizontal)

        # 3) 배정 테이블
        assign_frame = self._make_frame("배정")
        self.tbl_assign = self._make_table(["사번", "담당자명", "배정갯수"])
        self.tbl_assign.setColumnWidth(0, 80)
        self.tbl_assign.setColumnWidth(1, 80)
        self.tbl_assign.setColumnWidth(2, 70)
        assign_frame.layout().addWidget(self.tbl_assign)
        bottom_splitter.addWidget(assign_frame)

        # 4) 처리상황 로그
        log_frame = self._make_frame("처리상황")
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setFont(QFont("D2Coding", 9))
        log_frame.layout().addWidget(self.txt_log)
        bottom_splitter.addWidget(log_frame)

        bottom_splitter.setSizes([280, 900])
        root_layout.addWidget(bottom_splitter, stretch=4)

        # ── 버튼 영역 ───────────────────────
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        self.btn_query  = QPushButton("DB 조회")
        self.btn_detail = QPushButton("상세 조회")
        self.btn_assign = QPushButton("DB 배정")
        self.btn_input  = QPushButton("DB 입력")

        for btn in [self.btn_query, self.btn_detail, self.btn_assign, self.btn_input]:
            btn.setFixedHeight(40)
            btn.setFont(QFont("D2Coding", 10, QFont.Bold))

        self.btn_query.setStyleSheet("background:#1565C0; color:white; border-radius:4px;")
        self.btn_detail.setStyleSheet("background:#2E7D32; color:white; border-radius:4px;")
        self.btn_assign.setStyleSheet("background:#E65100; color:white; border-radius:4px;")
        self.btn_input.setStyleSheet("background:#B71C1C; color:white; border-radius:4px;")

        self.btn_query.clicked.connect(self.on_query)
        self.btn_detail.clicked.connect(self.on_detail)
        self.btn_assign.clicked.connect(self.on_assign)
        self.btn_input.clicked.connect(self.on_input)

        btn_layout.addWidget(self.btn_query)
        btn_layout.addWidget(self.btn_detail)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_assign)
        btn_layout.addWidget(self.btn_input)

        root_layout.addLayout(btn_layout)

    def _make_frame(self, title: str) -> QFrame:
        frame = QFrame()
        frame.setFrameShape(QFrame.StyledPanel)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(4, 4, 4, 4)
        layout.setSpacing(2)
        lbl = QLabel(f" {title}")
        lbl.setFont(QFont("D2Coding", 9, QFont.Bold))
        lbl.setStyleSheet("color:#90CAF9;")
        layout.addWidget(lbl)
        return frame

    def _make_table(self, headers: list) -> QTableWidget:
        tbl = QTableWidget(0, len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        tbl.setAlternatingRowColors(True)
        tbl.verticalHeader().setDefaultSectionSize(22)
        tbl.verticalHeader().setVisible(False)
        tbl.setFont(QFont("D2Coding", 9))
        return tbl

    def _apply_style(self):
        self.setStyleSheet("""
            QMainWindow, QWidget { background:#1B2631; color:#ECF0F1; }
            QTableWidget {
                background:#1E2D3D; alternate-background-color:#243447;
                gridline-color:#2C3E50; color:#ECF0F1;
                selection-background-color:#1565C0;
            }
            QHeaderView::section {
                background:#0D47A1; color:white; font-weight:bold;
                padding:3px; border:1px solid #1565C0;
            }
            QFrame { border:1px solid #2C3E50; border-radius:4px; }
            QTextEdit { background:#0D1117; color:#90EE90; border:none; }
            QScrollBar:vertical { background:#1B2631; width:10px; }
            QScrollBar::handle:vertical { background:#2C3E50; border-radius:4px; }
        """)

    # ── 로그 출력 ────────────────────────────
    def log(self, msg: str, color: str = ""):
        ts = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss")
        if color:
            html = f'<span style="color:{color}">{ts}&nbsp;&nbsp;{msg}</span>'
        else:
            html = f'<span style="color:#90EE90">{ts}&nbsp;&nbsp;{msg}</span>'
        self.txt_log.append(html)

    def log_error(self, msg: str):
        self.log(f"[오류] {msg}", "#FF5252")

    # ── Config 로드 ──────────────────────────
    def load_config(self):
        self.log("데이터베이스 설정 파일 확인 중...")
        db_path = os.path.join(DB_DIR, DB_FILE)
        if not os.path.exists(db_path):
            self.log("Config_DB.db 없음 → Google Drive에서 다운로드 중...")
            success, result = download_db()
            if success:
                self.log(f"✓ 다운로드 완료: {result}")
            else:
                self.log_error(f"다운로드 실패: {result}")
                QMessageBox.critical(self, "설정 로드 실패",
                    f"Config_DB.db 를 다운로드할 수 없습니다.\n\n오류: {result}")
                return
        else:
            self.log(f"✓ Config_DB.db 확인됨: {db_path}")
        try:
            self.db_config = load_db_config()
            self.log(f"✓ DB 타입      : {self.db_config['DB_Type']}")
            self.log(f"✓ 서버         : {self.db_config['Host']}:{self.db_config['Port']}")
            self.log(f"✓ 데이터베이스 : {self.db_config['DB_Name']}")
            self.log("✓ 설정 로드 완료")
        except Exception as e:
            self.log_error(str(e))
            QMessageBox.critical(self, "설정 로드 실패", str(e))

    # ─────────────────────────────────────────
    #  DB 조회 버튼
    # ─────────────────────────────────────────
    def on_query(self):
        if not self.db_config:
            QMessageBox.warning(self, "경고", "DB 설정이 로드되지 않았습니다.")
            return
        self.log("월별 DB 수량 조회 중...")
        self.btn_query.setEnabled(False)

        def task():
            sql = """
                SELECT 		
                DISTINCT LEFT(t1.Rec_Date,7) AS Rec_YM, count(OrderNo) CNT
                FROM dbo.TM_MEMBER t1 WITH(NOLOCK)		
                LEFT JOIN dbo.Member m1 WITH(NOLOCK) ON m1.MemberNo = t1.Mem_ID		
                LEFT JOIN dbo.Staff s1 WITH(NOLOCK) ON t1.AssignCharge_ID = s1.SaBun 		
                LEFT JOIN dbo.Staff s2 WITH(NOLOCK) ON m1.Charge_IDP = s2.SaBun		
                LEFT JOIN dbo.DoNotCall d1 WITH(NOLOCK) ON REPLACE(t1.Mobile,'-','') = d1.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(Mobile,'-','') AS Mobile FROM Member WITH(NOLOCK) WHERE MemType IN ('만기','정상','행사')) m2 ON REPLACE(t1.Mobile,'-','') = m2.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(Tel,'-','') AS Mobile FROM Member WITH(NOLOCK) WHERE MemType IN ('만기','정상','행사')) m3 ON REPLACE(t1.Mobile,'-','') = m3.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(OfficeTel,'-','') AS Mobile FROM Member WITH(NOLOCK) WHERE MemType IN ('만기','정상','행사')) m4 ON REPLACE(t1.Mobile,'-','') = m4.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(t1.Mobile,'-', '') AS Mobile FROM TM_MEMBER t1 WITH(NOLOCK) LEFT JOIN dbo.Staff s1 WITH(NOLOCK) ON t1.AssignCharge_ID = s1.SaBun WHERE s1.PlaceofDuty IN ('글로벌세무금융','올리고생활건강','미채결','바른라이프')) t2  ON t2.Mobile = t1.Mobile		
                WHERE HS_Name LIKE '%쇼핑%' AND Mem_ID = ''		
                AND d1.Mobile IS NULL -- 두낫콜 제외		
                AND s1.PlaceofDuty NOT IN ('글로벌세무금융', '올리고생활건강','미채결','바른라이프') -- 미체결 업체에서 기배정된 DB제외		
                AND REPLACE(t1.Mobile,'-','') NOT IN ('01083369251', '01083369251') -- CJ MD 및 결번 제외		
                AND m2.Mobile IS NULL -- 현재 유지 고객 번호 제외		
                AND m3.Mobile IS NULL -- 현재 유지 고객 번호 제외		
                AND m4.Mobile IS NULL -- 현재 유지 고객 번호 제외		
                AND t2.Mobile IS NULL -- 조아유에 배정된 번호 제외		
                GROUP BY LEFT(t1.Rec_Date,7)
                ORDER BY LEFT(t1.Rec_Date,7)
            """
            conn = get_mssql_connection(self.db_config)
            cursor = conn.cursor()
            cursor.execute(sql)
            rows = cursor.fetchall()
            conn.close()
            return rows

        self._worker = Worker(task)
        self._worker.result.connect(self._on_query_result)
        self._worker.error.connect(self._on_worker_error)
        self._worker.finished.connect(lambda: self.btn_query.setEnabled(True))
        self._worker.start()

    def _on_query_result(self, rows):
        self.tbl_monthly.setRowCount(0)
        for rec_ym, cnt in rows:
            row = self.tbl_monthly.rowCount()
            self.tbl_monthly.insertRow(row)
            # 체크박스
            chk = QCheckBox()
            chk.setStyleSheet("margin-left:6px;")
            self.tbl_monthly.setCellWidget(row, 0, chk)
            self.tbl_monthly.setItem(row, 1, self._item(str(rec_ym)))
            self.tbl_monthly.setItem(row, 2, self._item(str(cnt), Qt.AlignRight | Qt.AlignVCenter))
        self.log(f"월별 DB 수량 조회 완료  ({len(rows)}건)")

    # ─────────────────────────────────────────
    #  상세 조회 버튼
    # ─────────────────────────────────────────
    def on_detail(self):
        if not self.db_config:
            QMessageBox.warning(self, "경고", "DB 설정이 로드되지 않았습니다.")
            return

        checked_months = self._get_checked_months()
        if not checked_months:
            QMessageBox.warning(self, "선택 없음", "[월]을 선택하세요")
            return

        self.log(f"상세 조회 시작 → 선택 월: {', '.join(checked_months)}")
        self.btn_detail.setEnabled(False)

        months_placeholder = ",".join([f"'{m}'" for m in checked_months])

        def task():
            sql = f"""
                SELECT 		
                DISTINCT REPLACE(t1.Mobile,'-','') AS Mobile, t1.ID, t1.Name, t1.Rec_Date, MemoType, AssignCharge_ID, s1.saname, s1.PlaceofDuty, AssignDate, OrderNo, HS_Name		
                FROM dbo.TM_MEMBER t1 WITH(NOLOCK)		
                LEFT JOIN dbo.Member m1 WITH(NOLOCK) ON m1.MemberNo = t1.Mem_ID		
                LEFT JOIN dbo.Staff s1 WITH(NOLOCK) ON t1.AssignCharge_ID = s1.SaBun 		
                LEFT JOIN dbo.Staff s2 WITH(NOLOCK) ON m1.Charge_IDP = s2.SaBun		
                LEFT JOIN dbo.DoNotCall d1 WITH(NOLOCK) ON REPLACE(t1.Mobile,'-','') = d1.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(Mobile,'-','') AS Mobile FROM Member WITH(NOLOCK) WHERE MemType IN ('만기','정상','행사')) m2 ON REPLACE(t1.Mobile,'-','') = m2.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(Tel,'-','') AS Mobile FROM Member WITH(NOLOCK) WHERE MemType IN ('만기','정상','행사')) m3 ON REPLACE(t1.Mobile,'-','') = m3.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(OfficeTel,'-','') AS Mobile FROM Member WITH(NOLOCK) WHERE MemType IN ('만기','정상','행사')) m4 ON REPLACE(t1.Mobile,'-','') = m4.Mobile		
                LEFT JOIN (SELECT DISTINCT REPLACE(t1.Mobile,'-', '') AS Mobile FROM TM_MEMBER t1 WITH(NOLOCK) LEFT JOIN dbo.Staff s1 WITH(NOLOCK) ON t1.AssignCharge_ID = s1.SaBun WHERE s1.PlaceofDuty IN ('글로벌세무금융','올리고생활건강','미채결','바른라이프')) t2  ON t2.Mobile = t1.Mobile		
                WHERE HS_Name LIKE '%쇼핑%' AND Mem_ID = ''		
                AND d1.Mobile IS NULL -- 두낫콜 제외		
                AND s1.PlaceofDuty NOT IN ('글로벌세무금융', '올리고생활건강','미채결','바른라이프') -- 미체결 업체에서 기배정된 DB제외		
                AND REPLACE(t1.Mobile,'-','') NOT IN ('01083369251', '01083369251') -- CJ MD 및 결번 제외		
                AND m2.Mobile IS NULL -- 현재 유지 고객 번호 제외		
                AND m3.Mobile IS NULL -- 현재 유지 고객 번호 제외		
                AND m4.Mobile IS NULL -- 현재 유지 고객 번호 제외		
                AND t2.Mobile IS NULL -- 조아유에 배정된 번호 제외		
                and LEFT(t1.Rec_Date,7) in ({months_placeholder})
                ORDER BY t1.Rec_Date, AssignDate, t1.ID	
            """
            conn = get_mssql_connection(self.db_config)
            cursor = conn.cursor()
            cursor.execute(sql)
            rows = cursor.fetchall()
            # 배정 담당자 조회
            sql2 = """
                SELECT SaBun, SaName FROM Staff
                WHERE PlaceofDuty IN ('홈쇼핑 TM','미채결')
                AND OutDate = '' AND BranchOffice IN ('TM1','TM2')
                ORDER BY PlaceofDuty DESC, SaName
            """
            cursor.execute(sql2)
            staff = cursor.fetchall()
            conn.close()
            return rows, staff

        self._worker2 = Worker(task)
        self._worker2.result.connect(self._on_detail_result)
        self._worker2.error.connect(self._on_worker_error)
        self._worker2.finished.connect(lambda: self.btn_detail.setEnabled(True))
        self._worker2.start()

    def _on_detail_result(self, data):
        rows, staff = data
        self.detail_data = rows

        # 새 조회 시작 → 입력 완료 플래그 초기화, DB 입력 버튼 재활성화
        self._input_done = False
        self.btn_input.setEnabled(True)
        self.btn_input.setStyleSheet("background:#B71C1C; color:white; border-radius:4px;")
        self.btn_input.setText("DB 입력")

        # DB 내용 테이블
        self.tbl_detail.setRowCount(0)
        for r in rows:
            row = self.tbl_detail.rowCount()
            self.tbl_detail.insertRow(row)
            for col, val in enumerate(r):
                self.tbl_detail.setItem(row, col, self._item(str(val) if val is not None else ""))
            # 신규배정일(11), 신규담당자(12) 빈 컬럼
            self.tbl_detail.setItem(row, 11, self._item(""))
            self.tbl_detail.setItem(row, 12, self._item(""))

        self.log(f"선택월 DB 상세조회 완료  ({len(rows)}건)")

        # 배정 테이블
        self.assign_data = staff
        self.tbl_assign.setRowCount(0)
        for sabun, saname in staff:
            row = self.tbl_assign.rowCount()
            self.tbl_assign.insertRow(row)
            self.tbl_assign.setItem(row, 0, self._item(str(sabun)))
            self.tbl_assign.setItem(row, 1, self._item(str(saname)))
            # 배정갯수 - 편집 가능
            qty_item = QTableWidgetItem("0")
            qty_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_assign.setItem(row, 2, qty_item)

        self.tbl_assign.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        self.log(f"배정담당자 조회 완료  ({len(staff)}명)")

    # ─────────────────────────────────────────
    #  DB 배정 버튼
    # ─────────────────────────────────────────
    def on_assign(self):
        # 배정갯수 검증
        assign_list = self._get_assign_list()
        if not assign_list:
            self.log_error("배정갯수를 입력해 주세요")
            QMessageBox.warning(self, "입력 오류", "배정갯수를 입력해 주세요")
            return

        total_assign = sum(q for _, _, q in assign_list)
        total_detail = self.tbl_detail.rowCount()

        if total_detail == 0:
            QMessageBox.warning(self, "경고", "상세 조회를 먼저 실행하세요.")
            return

        if total_assign > total_detail:
            QMessageBox.warning(
                self, "수량 초과",
                f"배정 총 갯수({total_assign})가 DB 건수({total_detail})를 초과합니다."
            )
            return

        today = datetime.date.today().strftime("%Y-%m-%d")

        # 라운드로빈 배정 순서 생성
        sequence = self._build_assign_sequence(assign_list)

        # tbl_detail 에 신규배정일/신규담당자 입력
        for i, (sabun, saname) in enumerate(sequence):
            self.tbl_detail.item(i, 11).setText(today)
            self.tbl_detail.item(i, 12).setText(saname)
            # 배정된 행 배경 강조
            for col in range(self.tbl_detail.columnCount()):
                item = self.tbl_detail.item(i, col)
                if item:
                    item.setBackground(QColor("#1A3A2A"))

        self.log(f"DB 배정 완료  (배정: {len(sequence)}건, 배정일: {today})")

    def _build_assign_sequence(self, assign_list: list) -> list:
        """라운드로빈 방식으로 배정 순서 생성"""
        remaining = [(sabun, saname, qty) for sabun, saname, qty in assign_list]
        sequence  = []
        while any(q > 0 for _, _, q in remaining):
            for idx, (sabun, saname, qty) in enumerate(remaining):
                if qty > 0:
                    sequence.append((sabun, saname))
                    remaining[idx] = (sabun, saname, qty - 1)
        return sequence

    # ─────────────────────────────────────────
    #  DB 입력 버튼
    # ─────────────────────────────────────────
    def on_input(self):
        if not self.db_config:
            QMessageBox.warning(self, "경고", "DB 설정이 로드되지 않았습니다.")
            return

        # ── 중복 실행 방지 ──────────────────
        if self._input_done:
            QMessageBox.warning(
                self, "중복 실행 방지",
                "이미 DB 입력이 완료된 작업입니다.\n\n"
                "새로운 배정 작업을 하려면\n"
                "[상세 조회] 버튼을 다시 실행하세요."
            )
            return

        # 신규배정 행 수집
        new_assigned = []
        for row in range(self.tbl_detail.rowCount()):
            assign_date = self.tbl_detail.item(row, 11).text().strip() if self.tbl_detail.item(row, 11) else ""
            new_name    = self.tbl_detail.item(row, 12).text().strip() if self.tbl_detail.item(row, 12) else ""
            if assign_date and new_name:
                mem_id   = self.tbl_detail.item(row, 1).text().strip() if self.tbl_detail.item(row, 1) else ""
                sabun = self._find_sabun_by_name(new_name)
                new_assigned.append({
                    "ID"         : str(mem_id),
                    "AssignDate" : str(assign_date),
                    "SaBun"      : str(sabun),
                    "SaName"     : str(new_name),
                })

        if not new_assigned:
            QMessageBox.warning(self, "경고", "배정된 데이터가 없습니다.\n[DB 배정]을 먼저 실행하세요.")
            return

        reply = QMessageBox.question(
            self, "DB 입력 확인",
            f"총 {len(new_assigned)}건을 DB에 입력하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        self.btn_input.setEnabled(False)
        self.log(f"DB 입력 시작  ({len(new_assigned)}건)...")

        def task():
            conn   = get_mssql_connection(self.db_config)
            cursor = conn.cursor()
            print(new_assigned)
            sql = """
                UPDATE TM_MEMBER
                SET AssignCharge_ID = %s, AssignDate = %s 
                WHERE ID = %s
            """
            sql2 = """
                
            """
            for rec in new_assigned:
                # print(sql, (rec["SaBun"], rec["AssignDate"], rec["ID"]))
                cursor.execute(sql, (rec["SaBun"], rec["AssignDate"], rec["ID"]))
                # print("OK")
                
            conn.commit()
            conn.close()
            return new_assigned

        self._worker3 = Worker(task)
        self._worker3.result.connect(self._on_input_result)
        self._worker3.error.connect(self._on_worker_error)
        # 오류 발생 시에만 버튼 재활성화 (정상 완료 시에는 _on_input_result에서 잠금)
        self._worker3.error.connect(lambda: self.btn_input.setEnabled(True))
        self._worker3.start()

    def _on_input_result(self, new_assigned):
        self.log(f"✓ DB 입력 완료  ({len(new_assigned)}건)")
        # 완료 플래그 세팅 + 버튼 완전 잠금
        self._input_done = True
        self.btn_input.setEnabled(False)
        self.btn_input.setStyleSheet(
            "background:#424242; color:#757575; border-radius:4px;"
        )
        self.btn_input.setText("DB 입력 (완료)")
        self._export_excel(new_assigned)

    # ─────────────────────────────────────────
    #  엑셀 내보내기
    # ─────────────────────────────────────────
    def _export_excel(self, new_assigned: list):
        today = datetime.date.today().strftime("%Y-%m-%d")
        download_dir = os.path.join(app_dir(), "Downloads")
        os.makedirs(download_dir, exist_ok=True)

        # ── 중복 파일명 처리: 동일 파일 존재 시 _1, _2 ... 순서로 증가 ──
        base_name = f"미체결_DB배정_{today}"
        save_path = os.path.join(download_dir, f"{base_name}.xlsx")
        counter = 1
        while os.path.exists(save_path):
            save_path = os.path.join(download_dir, f"{base_name}_{counter}.xlsx")
            counter += 1

        headers = [
            "휴대전화", "ID", "성명", "등록일", "상태",
            "사번", "담당자명", "권", "배정일",
            "주문번호", "채널", "신규배정일", "신규담당자"
        ]

        # detail 테이블에서 신규배정된 행만 추출
        assigned_ids = {r["ID"] for r in new_assigned}
        data_rows = []
        for row in range(self.tbl_detail.rowCount()):
            mem_id = self.tbl_detail.item(row, 1).text() if self.tbl_detail.item(row, 1) else ""
            if mem_id in assigned_ids:
                row_data = []
                for col in range(self.tbl_detail.columnCount()):
                    item = self.tbl_detail.item(row, col)
                    row_data.append(item.text() if item else "")
                data_rows.append(row_data)

        wb = openpyxl.Workbook()

        # ── 전체 시트 ──
        ws_all = wb.active
        ws_all.title = "전체"
        self._write_sheet(ws_all, headers, data_rows)

        # ── 담당자별 시트 ──
        from collections import defaultdict
        by_person = defaultdict(list)
        for row_data in data_rows:
            name = row_data[12]  # 신규담당자
            by_person[name].append(row_data)

        for person, rows in sorted(by_person.items()):
            ws = wb.create_sheet(title=person[:31])  # 시트명 31자 제한
            self._write_sheet(ws, headers, rows)

        wb.save(save_path)
        self.log(f"✓ 엑셀 저장 완료: {save_path}")
        QMessageBox.information(self, "엑셀 저장", f"파일이 저장되었습니다:\n{save_path}")

    def _write_sheet(self, ws, headers: list, data_rows: list):
        # 헤더 스타일
        hdr_fill = PatternFill("solid", fgColor="0D47A1")
        hdr_font = XLFont(bold=True, color="FFFFFF", name="맑은 고딕", size=9)
        thin = Side(style="thin", color="BBBBBB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

        # 데이터
        data_font = XLFont(name="맑은 고딕", size=9)
        for r_idx, row_data in enumerate(data_rows, 2):
            fill_color = "EBF5FB" if r_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill("solid", fgColor=fill_color)
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = data_font
                cell.fill      = row_fill
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

        # 열 너비 자동 조정
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ─────────────────────────────────────────
    #  헬퍼 메서드
    # ─────────────────────────────────────────
    def _item(self, text: str, align=Qt.AlignLeft | Qt.AlignVCenter) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setTextAlignment(align)
        return item

    def _get_checked_months(self) -> list:
        months = []
        for row in range(self.tbl_monthly.rowCount()):
            widget = self.tbl_monthly.cellWidget(row, 0)
            if widget and isinstance(widget, QCheckBox) and widget.isChecked():
                item = self.tbl_monthly.item(row, 1)
                if item:
                    months.append(item.text().strip())
        return months

    def _get_assign_list(self) -> list:
        result = []
        for row in range(self.tbl_assign.rowCount()):
            sabun  = self.tbl_assign.item(row, 0).text().strip() if self.tbl_assign.item(row, 0) else ""
            saname = self.tbl_assign.item(row, 1).text().strip() if self.tbl_assign.item(row, 1) else ""
            qty_text = self.tbl_assign.item(row, 2).text().strip() if self.tbl_assign.item(row, 2) else "0"
            try:
                qty = int(qty_text)
            except ValueError:
                qty = 0
            if qty > 0:
                result.append((sabun, saname, qty))
        return result

    def _find_sabun_by_name(self, saname: str) -> str:
        for row in range(self.tbl_assign.rowCount()):
            name = self.tbl_assign.item(row, 1).text().strip() if self.tbl_assign.item(row, 1) else ""
            if name == saname:
                sabun = self.tbl_assign.item(row, 0).text().strip() if self.tbl_assign.item(row, 0) else ""
                return sabun
        return ""

    def _on_worker_error(self, msg: str):
        self.log_error(msg)
        QMessageBox.critical(self, "오류", msg)


# ─────────────────────────────────────────────
#  진입점
# ─────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("미체결DB 배정")

    # ── D2Coding 폰트 로드 ──────────────────────
    font_path = resource_path(os.path.join("Font", "D2Coding-Ver1.3.2-20180524.ttf"))
    if os.path.exists(font_path):
        QFontDatabase.addApplicationFont(font_path)
        app.setFont(QFont("D2Coding", 10))
    # ───────────────────────────────────────────
    window = MainWindow()
    window.show()
    sys.exit(app.exec())